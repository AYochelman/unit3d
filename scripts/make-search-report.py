# -*- coding: utf-8 -*-
"""Builds the "what people search for" workbook for the shop.

Run:  python scripts/make-search-report.py
Out:  Unit3D-Search-Categories.xlsx  (in the repo root)

The demand numbers are RELATIVE (1-5), based on what consistently trends on the
maker platforms and on Israeli gift/e-commerce search behaviour - they are a
prioritisation aid, not measured search volume. The model column holds SEARCH
TERMS plus a ready search URL rather than direct model links, so nothing here
is a guessed URL.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GREEN = "089A47"; DARK = "1C1C1F"; LIGHT = "F2F2F4"; AMBER = "FFB02E"

ROWS = [
    # (קטגוריה, למה, ביקוש, תחרות, עמודה באתר, טווח מחיר, זמן הדפסה, עדיפות, מונחי חיפוש)
    ("דרקונים מפרקיים (Flexi)", "המוצר שהכי מזוהה עם הדפסת תלת מימד. ויראלי בטיקטוק, מתנה מנצחת לילדים ולמבוגרים.", 5, 5, "פידג'טים › פלקסי", "90-160", "4-9h", 1, "flexi dragon; articulated dragon; crystal dragon"),
    ("חיות מפרקיות אחרות", "נחשים, תמנונים, אקסולוטלים, כרישים. אותו ביקוש, פחות תחרות מדרקונים.", 5, 4, "פידג'טים › פלקסי", "70-140", "3-8h", 1, "flexi axolotl; articulated snake; flexi octopus; flexi shark"),
    ("פידג'טים לשולחן", "ספינרים, קוביות אינסוף, סליידרים. מחיר כניסה נמוך, קונים כמה ביחד.", 4, 5, "פידג'טים › פידג'ט", "45-80", "1-2h", 2, "fidget slider; infinity cube; fidget spinner; click fidget"),
    ("תגי שם לחיות", "חיפוש עם כוונת קנייה גבוהה. חוזר כל כמה חודשים (התג הולך לאיבוד).", 4, 3, "תגים לחיות", "30-45", "25-45min", 1, "pet tag; dog tag name; cat id tag"),
    ("מתנות מותאמות אישית עם שם", "החיפוש הכי מסחרי בעברית. שלטים, מחזיקים, סימניות עם שם.", 5, 3, "מעצב אישי", "30-70", "0.5-2h", 1, "name keychain; personalized nameplate; custom name sign"),
    ("סמלי יחידות צה\"ל", "ייחודי לשוק הישראלי, כמעט בלי תחרות מודפסת. שיא בעונת הטקסים.", 4, 1, "סמלי יחידה", "55-120", "1.5-3.5h", 1, "military badge; unit emblem; army insignia"),
    ("קייסים לטלפון", "ביקוש ענק, אבל דורש דגם מדויק. TPU בלבד.", 4, 5, "מעצב אישי", "70-110", "2-3h", 3, "phone case tpu; iphone case print"),
    ("בוסטים ופסלי דיוקן", "מחיר גבוה ליחידה, שולי רווח טובים. מתאים למתנה יוקרתית.", 3, 3, "פסלים", "230-450", "8-16h", 2, "bust statue; portrait bust; sculpture bust"),
    ("חיות לואו-פולי", "אסתטיקה שמוכרת את עצמה בתמונה אחת. מהיר יחסית לפסל.", 4, 3, "פסלים", "120-200", "5-8h", 2, "low poly animal; low poly deer; low poly wolf"),
    ("אגרטלים ו-Vase Mode", "הדפסה בקיר אחד: מעט חומר, תוצאה מרשימה. מרווח מצוין.", 4, 3, "פסלים", "90-160", "4-6h", 2, "spiral vase; vase mode; twisted vase"),
    ("מנורות ואור", "מנורת ירח, מנורות ליתופן. מוצר מתנה מובהק.", 4, 3, "פסלים", "140-220", "7-10h", 2, "moon lamp; lithophane lamp; night light print"),
    ("גביעים ופרסים", "B2B: טורנירים, מועדונים, \"עובד החודש\". הזמנות בכמות.", 3, 2, "פסלים / עסקים", "150-260", "6-10h", 2, "trophy award; custom trophy; award statue"),
    ("משחקי שולחן ואביזרים", "מגדלי קוביות, מחזיקי קלפים, אורגנייזרים למשחקים. קהל נאמן.", 4, 3, "לבית ולמשרד", "50-120", "2-5h", 2, "dice tower; card holder board game; token tray"),
    ("שחמט ומשחקי אסטרטגיה", "סט שלם = מחיר גבוה. גם מתנה וגם פריט תצוגה.", 3, 3, "פסלים", "350-500", "18-25h", 3, "chess set print; chess pieces 3d"),
    ("מארגנים לשולחן ולמגירה", "Gridfinity ודומיו. קונים 10 יחידות בבת אחת.", 4, 4, "לבית ולמשרד", "35-90", "1.5-4h", 2, "gridfinity; drawer organizer; desk organizer"),
    ("מעמדים לטלפון ולאוזניות", "מוצר כניסה זול, נמכר טוב לגיימרים.", 4, 5, "לבית ולמשרד", "35-60", "1-2.5h", 3, "phone stand; headphone stand; controller holder"),
    ("עציצים ואדניות", "עונתי (אביב), נמכר טוב באינסטגרם.", 3, 4, "לבית ולמשרד", "45-90", "2.5-4h", 3, "geometric planter; succulent pot; self watering planter"),
    ("חלקי חילוף ומתאמים", "מרווח גבוה, אפס תחרות. הלקוח מגיע עם צורך אמיתי.", 3, 1, "העלאת קובץ", "לפי הצעה", "0.5-3h", 2, "replacement part; bracket adapter; clip repair"),
    ("קישוטי חג ועונה", "ראש השנה, חנוכה, פורים. שיא חד לכמה שבועות.", 4, 3, "טרנדי כרגע", "30-80", "1-3h", 2, "hanukkah decor; holiday ornament; menorah print"),
    ("מתנות ממותגות לעסקים", "ההזמנה הגדולה ביותר לפי לקוח. 10-100 יחידות בהזמנה.", 3, 2, "לעסקים", "35-90 ליח'", "0.5-2h", 1, "corporate gift print; logo keychain; branded desk item"),
]

HEAD = ["קטגוריה", "למה כדאי", "ביקוש (1-5)", "תחרות (1-5)", "לאיזו עמודה באתר",
        "טווח מחיר ₪", "זמן הדפסה", "עדיפות", "מונחי חיפוש למייקרוורלד", "קישור חיפוש מוכן"]

def style_header(ws, ncols):
    for i in range(1, ncols + 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "A2"

def main():
    wb = Workbook()
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1 ──────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "קטגוריות מומלצות"
    ws.sheet_view.rightToLeft = True
    ws.append(HEAD)
    for r in ROWS:
        term = r[8]
        url = "https://makerworld.com/en/search/models?keyword=" + term.split(";")[0].strip().replace(" ", "+")
        ws.append(list(r) + [url])
    style_header(ws, len(HEAD))

    widths = [30, 52, 12, 12, 22, 14, 14, 10, 46, 46]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(HEAD)):
        for c in row:
            c.border = border
            c.alignment = Alignment(vertical="top", wrap_text=True)
        prio = row[7].value
        if prio == 1:
            for c in row:
                c.fill = PatternFill("solid", fgColor="E8F5EC")
        row[2].alignment = Alignment(horizontal="center", vertical="center")
        row[3].alignment = Alignment(horizontal="center", vertical="center")
        row[7].alignment = Alignment(horizontal="center", vertical="center")
        row[7].font = Font(bold=True)
        row[9].font = Font(color="0563C1", underline="single", size=9)
        row[9].hyperlink = row[9].value

    # ── Sheet 2 ──────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("איך ממלאים מודלים")
    ws2.sheet_view.rightToLeft = True
    ws2.append(["שלב", "מה עושים"])
    steps = [
        ("1", "פותחים את הגיליון הראשון ומסתכלים על העמודה 'עדיפות'. מתחילים מהשורות שמסומנות 1."),
        ("2", "לוחצים על הקישור בעמודה האחרונה. הוא פותח את החיפוש במייקרוורלד עם המונח המתאים."),
        ("3", "בוחרים את המודלים שאהבתם ומוסיפים אותם לאוסף במייקרוורלד (Collection)."),
        ("4", "מעתיקים את כתובת האוסף ומוסיפים אותה לקובץ scripts/makerworld-sources.json בפרויקט."),
        ("5", "לוחצים פעמיים על import-models.bat. הסקריפט מוריד את הרשימה ומכניס אותה לאתר."),
        ("6", "מריצים npm run build ומעלים. המודלים מופיעים בעמודה שהוגדרה להם."),
        ("", ""),
        ("שים לב", "מודלים ברישיון NC (לא-מסחרי) מיובאים אבל מוסתרים מהחנות, כי אסור למכור הדפסה שלהם בלי אישור מהמעצב."),
        ("שים לב", "לכל מודל מיובא נשמר שם המעצב וקישור למקור - זו הדרישה של רישיון CC-BY."),
    ]
    for s in steps:
        ws2.append(list(s))
    style_header(ws2, 2)
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 110
    for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row, max_col=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
        if row[0].value == "שים לב":
            for c in row:
                c.fill = PatternFill("solid", fgColor="FFF4E0")
            row[0].font = Font(bold=True, color="9A6700")

    # ── Sheet 3 ──────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("הסבר על המספרים")
    ws3.sheet_view.rightToLeft = True
    ws3.append(["עמודה", "מה זה אומר"])
    notes = [
        ("ביקוש 1-5", "כמה אנשים מחפשים את זה. 5 = חיפוש רחב ויציב לאורך השנה."),
        ("תחרות 1-5", "כמה קל למצוא את זה כבר מודפס ונמכר בארץ. 5 = הרבה מוכרים, 1 = כמעט אף אחד."),
        ("עדיפות", "1 = להתחיל מזה. 2 = בהמשך. 3 = רק אם נשאר זמן."),
        ("טווח מחיר", "מחיר מכירה מומלץ בשקלים, מבוסס על מודל העלות של האתר (חומר + מכונה + חשמל + עבודה) עם רווח יעד של 60%."),
        ("המספרים", "אלה הערכות יחסיות לתעדוף, לא נתוני חיפוש מדודים. אחרי חודש מכירות החליפו אותם בנתונים שלכם."),
    ]
    for n in notes:
        ws3.append(list(n))
    style_header(ws3, 2)
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 100
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row, max_col=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)

    out = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Unit3D-Search-Categories.xlsx")
    wb.save(out)
    print("wrote", out)

if __name__ == "__main__":
    main()
